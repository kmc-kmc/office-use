import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Define the target date for comparison
target_date = datetime.strptime('1/3/2024', "%d/%m/%Y").date()

# Upload the Excel file
from google.colab import files
uploaded = files.upload()

# Load the Excel file
file_name = list(uploaded.keys())[0]  # Get the uploaded file name
workbook = load_workbook(file_name)

# Initialize variables to store row information
definedrow = None

# Loop through each sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Step 1: Search for the cell with target date and save the row
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col).value

            # Convert cell value to a date if possible
            if isinstance(cell, str):
                try:
                    cell_date = datetime.strptime(cell, "%d/%m/%Y").date()
                except ValueError:
                    continue
            elif isinstance(cell, datetime):
                cell_date = cell.date()
            else:
                continue

            # Check if the cell date matches the target date
            if cell_date == target_date:
                definedrow = row  # Save the row number
                break  # Exit the loop after finding the target date
        if definedrow is not None:
            break

    # Step 2: Check from column B of definedrow
    if definedrow is not None:
        last_cell_value = sheet.cell(row=definedrow, column=sheet.max_column).value

        # Find the first cell with a value starting from column B
        first_cell_column = None
        for col in range(2, sheet.max_column + 1):  # Start from column B (2)
            cell_value = sheet.cell(row=definedrow, column=col).value
            if cell_value is not None:
                first_cell_column = col  # Save the column index of the first non-empty cell
                break  # Exit the loop after finding the first non-empty cell

        # Step 3: Copy the value of the last cell to the first cell with value
        if first_cell_column is not None:
            # Copy last cell value to every cell from the first non-empty cell to the last column
            for col in range(first_cell_column, sheet.max_column + 1):
                sheet.cell(row=definedrow, column=col).value = last_cell_value

# Save the modified workbook
new_file_name = 'modified_' + file_name
workbook.save(new_file_name)

# Download the modified workbook
files.download(new_file_name)
