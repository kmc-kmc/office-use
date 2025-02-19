import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from google.colab import files

# Define the target date for comparison
target_date = datetime.strptime('1/12/2023', "%d/%m/%Y").date()

# Upload the Excel file
uploaded = files.upload()

# Load the Excel file
file_name = list(uploaded.keys())[0]  # Get the uploaded file name
workbook = load_workbook(file_name)

# Loop through each sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Step 1: Search for the cell with "1/12/2023" and save the row
    definedrow = None
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col).value

            # Convert cell value to a date if possible
            if isinstance(cell, str):
                try:
                    cell_date = datetime.strptime(cell, "%d/%m/%Y").date()
                except ValueError:
                    continue
            elif isinstance(cell, datetime):
                cell_date = cell.date()
            else:
                continue

            # Check if the cell date matches the target date
            if cell_date == target_date:
                definedrow = row  # Save the row number
                break  # Exit the loop after finding the target date
        if definedrow is not None:
            break

    # Step 2: Clear all values in the defined row
    if definedrow is not None:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=definedrow, column=col).value = None

# Save the modified workbook
new_file_name = 'modified_' + file_name
workbook.save(new_file_name)

# Download the modified workbook
files.download(new_file_name)

# Output the defined row for verification
print(f'Defined row with the target date "1/12/2023": {definedrow}')
